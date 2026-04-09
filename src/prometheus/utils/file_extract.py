"""Shared file text extraction for Telegram gateway + Beacon web uploads.

Both gateways call extract_text() with a file path and optional MIME type.
Supported formats:
  - .txt, .md, .csv, .json, .yaml, .toml, .py, .js, .ts, .sh, .sql, etc. → UTF-8 read
  - .pdf → PyMuPDF (fitz)
  - .docx → python-docx
  - .xlsx → openpyxl
"""

from __future__ import annotations

import logging
import os
from pathlib import Path

logger = logging.getLogger(__name__)

MAX_EXTRACT_BYTES = 150_000  # ~150 KB text limit to avoid blowing up context

# Text-readable extensions (read as UTF-8)
TEXT_EXTENSIONS = {
    ".txt", ".md", ".csv", ".json", ".yaml", ".yml", ".toml",
    ".py", ".js", ".ts", ".tsx", ".jsx", ".sh", ".sql",
    ".html", ".xml", ".log", ".ini", ".cfg", ".env",
    ".dockerfile", ".makefile", ".rs", ".go", ".java", ".c", ".cpp", ".h",
    ".swift", ".kt", ".rb", ".pl", ".lua", ".r", ".m",
}

# Binary formats with specialized extractors
BINARY_EXTRACTORS = {".pdf", ".docx", ".xlsx"}

SUPPORTED_EXTENSIONS = TEXT_EXTENSIONS | BINARY_EXTRACTORS

UNSUPPORTED_MSG = (
    "I can't read {ext} files yet. "
    "Supported formats: txt, md, csv, json, yaml, py, pdf, docx, xlsx "
    "(and other common text/code files)."
)


def extract_text(filepath: str, mime_type: str | None = None) -> str | None:
    """Extract text content from a file. Returns None if unsupported or empty.

    This is the single entry point — both Telegram and Beacon call this.
    """
    path = Path(filepath)
    if not path.is_file():
        logger.warning("extract_text: file not found: %s", filepath)
        return None

    ext = path.suffix.lower()

    # Size guard
    size = path.stat().st_size
    if size == 0:
        return None

    # Text-based files
    if ext in TEXT_EXTENSIONS:
        return _extract_plain_text(path)

    # PDF
    if ext == ".pdf":
        return _extract_pdf(path)

    # DOCX
    if ext == ".docx":
        return _extract_docx(path)

    # XLSX
    if ext == ".xlsx":
        return _extract_xlsx(path)

    return None


def is_supported(filepath: str) -> bool:
    """Check if a file extension is supported for text extraction."""
    return Path(filepath).suffix.lower() in SUPPORTED_EXTENSIONS


def unsupported_message(filepath: str) -> str:
    """Return a user-friendly message for unsupported file types."""
    ext = Path(filepath).suffix.lower()
    return UNSUPPORTED_MSG.format(ext=ext or "this type of")


def _extract_plain_text(path: Path) -> str | None:
    """Read a text file as UTF-8."""
    try:
        data = path.read_bytes()
        if len(data) > MAX_EXTRACT_BYTES:
            text = data[:MAX_EXTRACT_BYTES].decode("utf-8", errors="replace")
            return text + f"\n\n[... truncated at {MAX_EXTRACT_BYTES // 1000} KB]"
        return data.decode("utf-8", errors="replace")
    except Exception as exc:
        logger.warning("Failed to read text file %s: %s", path, exc)
        return None


def _extract_pdf(path: Path) -> str | None:
    """Extract text from PDF using PyMuPDF (fitz)."""
    try:
        import fitz  # PyMuPDF
    except ImportError:
        logger.warning("PyMuPDF not installed — cannot extract PDF text. Install: uv add PyMuPDF")
        return f"[PDF file: {path.name} — install PyMuPDF to extract text]"

    try:
        doc = fitz.open(str(path))
        pages = []
        total_chars = 0
        for page in doc:
            text = page.get_text()
            if total_chars + len(text) > MAX_EXTRACT_BYTES:
                pages.append(text[: MAX_EXTRACT_BYTES - total_chars])
                pages.append(f"\n\n[... truncated at page {page.number + 1}/{len(doc)}]")
                break
            pages.append(text)
            total_chars += len(text)
        doc.close()
        result = "\n".join(pages).strip()
        return result if result else f"[PDF file: {path.name} — no extractable text (may be image-based)]"
    except Exception as exc:
        logger.warning("Failed to extract PDF %s: %s", path, exc)
        return f"[PDF file: {path.name} — extraction failed: {exc}]"


def _extract_docx(path: Path) -> str | None:
    """Extract text from DOCX using python-docx."""
    try:
        from docx import Document
    except ImportError:
        logger.warning("python-docx not installed — cannot extract DOCX text. Install: uv add python-docx")
        return f"[DOCX file: {path.name} — install python-docx to extract text]"

    try:
        doc = Document(str(path))
        paragraphs = []
        total_chars = 0
        for para in doc.paragraphs:
            text = para.text
            if total_chars + len(text) > MAX_EXTRACT_BYTES:
                paragraphs.append(f"\n[... truncated at {MAX_EXTRACT_BYTES // 1000} KB]")
                break
            paragraphs.append(text)
            total_chars += len(text)
        result = "\n".join(paragraphs).strip()
        return result if result else f"[DOCX file: {path.name} — no text content]"
    except Exception as exc:
        logger.warning("Failed to extract DOCX %s: %s", path, exc)
        return f"[DOCX file: {path.name} — extraction failed: {exc}]"


def _extract_xlsx(path: Path) -> str | None:
    """Extract text from XLSX using openpyxl."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.warning("openpyxl not installed — cannot extract XLSX text. Install: uv add openpyxl")
        return f"[XLSX file: {path.name} — install openpyxl to extract text]"

    try:
        wb = load_workbook(str(path), read_only=True, data_only=True)
        sheets = []
        total_chars = 0
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            rows.append(f"## Sheet: {sheet_name}")
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                line = " | ".join(cells)
                if total_chars + len(line) > MAX_EXTRACT_BYTES:
                    rows.append(f"[... truncated at {MAX_EXTRACT_BYTES // 1000} KB]")
                    total_chars = MAX_EXTRACT_BYTES + 1
                    break
                rows.append(line)
                total_chars += len(line) + 1
            sheets.append("\n".join(rows))
            if total_chars > MAX_EXTRACT_BYTES:
                break
        wb.close()
        result = "\n\n".join(sheets).strip()
        return result if result else f"[XLSX file: {path.name} — empty spreadsheet]"
    except Exception as exc:
        logger.warning("Failed to extract XLSX %s: %s", path, exc)
        return f"[XLSX file: {path.name} — extraction failed: {exc}]"
